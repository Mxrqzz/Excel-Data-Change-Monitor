from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

planilha = load_workbook('Dados PDP.xlsx')
dados = planilha.active

# Data Regra
data_regra = datetime(2024, 8, 13, 0, 0)

# Definindo cores de preenchimento
verde = PatternFill(
    start_color="11AD45", end_color="11AD45", fill_type="solid")

amarelo = PatternFill(
    start_color="E8E520", end_color="E8E520", fill_type="solid")

vermelho = PatternFill(
    start_color="00FF0000", end_color="00FF0000", fill_type="solid")

# Alterando Cor da linha se a informação tiver sido criada a partir do dia 13/08

# Iterando pelas linhas
for criacao in dados.iter_rows(min_row=2, min_col=33, max_col=35):
    criado = criacao[0].value
    modificado = criacao[2].value

    # Verifica se criado e modificado estão com formato de datetime
    if isinstance(criado, datetime) and isinstance(modificado, datetime):
        # Verifica se criado e modificado são maiores ou igual a data regra;
        if criado >= data_regra and criado == modificado:
            linha_atual = criacao[0].row
            for linha in dados[linha_atual]:
                linha.fill = verde
        elif modificado >= data_regra and criado < data_regra:
            linha_atual = criacao[0].row
            for linha in dados[linha_atual]:
                linha.fill = amarelo

# Alterando Cor da linha se a informação tiver sido excluida a partir do dia 13/08

# Iterando pelas linhas
for acao_validada in dados.iter_rows(min_row=2, min_col=32, max_col=35):
    acao = acao_validada[0].value
    modificado = acao_validada[3].value

    if acao == False and modificado >= data_regra:
        linha_atual = acao_validada[0].row
        for linha in dados[linha_atual]:
            linha.fill = vermelho

print("Classificao concluida")

planilha.save("PDP revisado.xlsx")

print("Planilha salva")